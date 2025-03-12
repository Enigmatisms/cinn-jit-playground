import re
import os
import sys
from pathlib import Path
from openpyxl import Workbook

METRIC_ORDER = [
    "Duration", 
    "Max Bandwidth", 
    "L1/TEX Hit Rate",
    "L2 Hit Rate",
    "Achieved Occupancy",
    "Theoretical Occupancy",
    "Registers Per Thread"
]

CALLBACKS = {
    "Duration": lambda x, y: x * 1000 if y == 'ms' else x,
    "Max Bandwidth": lambda x, _: x,
    "L1/TEX Hit Rate": lambda x, _: x,
    "L2 Hit Rate": lambda x, _: x,
    "Achieved Occupancy": lambda x, _: x,
    "Theoretical Occupancy": lambda x, _: x,
    "Registers Per Thread": lambda x, _: x
}

def parse_element(parts, callback = lambda x, y: x):

    if len(parts) > 1:
        value = float(parts[-1])  # The second last part is the value (number)
        unit = parts[-2]          # The last part is the unit (e.g., 'us')
        
        # Call the callback function with the value and unit
        return (callback(value, unit), True)
    return (0, False)

def parse_kernel_log(file_path, kernel_prefix):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    kernel_start_regex = re.compile(r'.*Context.*Stream.*Device.*')
    kernels = []
    current_kernel = None

    for i, line in enumerate(lines):
        # Check if the line contains the required words and the kernel prefix
        if kernel_start_regex.match(line):
            if current_kernel:
                kernels.append(current_kernel)
            if kernel_prefix in line:
                # If a kernel was previously found, save it
                if current_kernel:
                    kernels.append(current_kernel)
                
                # Start a new kernel
                current_kernel = {'start_line': i, 'name': line.strip().split(' ')[0], 'lines': [line.strip()]}
            else:
                current_kernel = None
        
        elif current_kernel:
            current_kernel['lines'].append(line.strip())
    
    # Don't forget to add the last kernel found (if any)
    if current_kernel:
        kernels.append(current_kernel)
    return kernels

def get_target_kernel_info(report_path: str, kernel_prefix: str):
    kernels = parse_kernel_log(report_path, kernel_prefix)
    if not kernels:
        print(f"File '{report_path}' has no valid kernel with prefix '{kernel_prefix}'")
        return [0 for _ in METRIC_ORDER]
    kernel = kernels[0]       # assume the first one is what we want
    # set metrics
    
    metrics_output = {
        "Duration": 0,
        "Max Bandwidth": 0,
        "L1/TEX Hit Rate": 0,
        "L2 Hit Rate": 0,
        "Achieved Occupancy": 0,
        "Theoretical Occupancy": 0,
        "Registers Per Thread": 0
    }
    
    for line in kernel["lines"]:
        elements = re.split(r'\s{2,}', line.strip())
        if elements:
            metric = elements[0]
            if metric in CALLBACKS:
                action = CALLBACKS[metric]
                output, flag = parse_element(elements, action)
                if not flag:
                    raise ValueError(f"'{metric}' parsing failed with flag as False. Please check what happened.")
                metrics_output[metric] = output
    
    result = []
    for metric in METRIC_ORDER:
        result.append(metrics_output[metric])
    
    return result

def batched_parsing(folder: str, kernel_prefix: str = 'fn_full_div_full_div_mul'):
    results = []
    for file in os.listdir(folder):
        file_path = os.path.join(folder, file)
        if os.path.isdir(file_path): continue
        data = get_target_kernel_info(file_path, kernel_prefix)
        elements = Path(file).stem.strip().split('-')
        data.insert(0, ','.join(elements[1:]))
        results.append(data)
    return results

def write_to_excel(headers, data, file_name='output.xlsx', column_width = 20):
    wb = Workbook()
    ws = wb.active
    
    for col in range(2, len(headers) + 2):      # set column width
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = column_width
    
    for i, header in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=header)
    
    for row_index, row_data in enumerate(data, start=2):
        # first column is index
        ws.cell(row=row_index, column=1, value=row_index - 1)
        # data, after the first column
        for col_index, value in enumerate(row_data, start=2):
            ws.cell(row=row_index, column=col_index, value=value)
    
    wb.save(file_name)

if __name__ == "__main__":
    # Example usage
    folder_path   = './data'
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <path to NCU cmdline report folder: ./data by default>")
    else:
        folder_path = sys.argv[1]
    kernel_prefix = 'fn_full_div_full'
    output_path   = './output.xlsx'
    
    all_data = batched_parsing(folder_path, kernel_prefix)
    
    headers = [
        "Shape (NHWC)",
        "Duration (us)", 
        "Max Bandwidth (%)", 
        "L1/TEX Hit Rate (%)",
        "L2 Hit Rate (%)",
        "Achieved Occupancy (%)",
        "Theoretical Occupancy (%)",
        "Registers Per Thread (#)"
    ]
    
    write_to_excel(headers, all_data, output_path)
